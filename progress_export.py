"""Builds Excel (.xlsx) and PDF exports of a Progress Tracker project.

The Excel export is the plain table only (Item / Item Description / Unit /
Suggested Quantity / % Project Cost / Week N Progress % / Week N Weighted %
/ ... / a TOTAL row) — no completion cards. Each "Week N Progress %" header
spans two columns (the raw entered % and the computed Week n/100 * %
Project Cost), mirroring the pair shown inline in the web table. Colors:
column headers dark grey, category rows dark gray, subcategory rows light
gray — all bold, along with the TOTAL row.

The PDF export additionally leads with the same "Overall Project
Completion" card and per-category cards shown on the web page, before the
table. Both cards mirror the web page's math exactly (see
computeGroupProgress / severityFor in progress_tracker.js): a group's
percent is the latest week's weighted-to-date cost divided by the group's
total % Project Cost — not summed across every week.
"""

from __future__ import annotations

import io

from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

LEVEL_CATEGORY = 0
LEVEL_SUBCATEGORY = 1
LEVEL_ITEM = 2

STATIC_HEADERS = ["Item", "Item Description", "Unit", "Suggested Quantity", "% Project Cost"]

HEADER_BG = "333333"
HEADER_FONT_COLOR = "FFFFFF"
CATEGORY_BG = "595959"
CATEGORY_FONT_COLOR = "FFFFFF"
SUBCATEGORY_BG = "D9D9D9"
SUBCATEGORY_FONT_COLOR = "000000"

# Mirrors the --success-text / --warning-text / --error-text / --text-muted
# and --border / --bg tones used by .pt-overall / .pt-card in style.css.
GOOD_COLOR = "1e7a34"
WARNING_COLOR = "b45309"
CRITICAL_COLOR = "c62828"
NONE_COLOR = "6b7280"
METER_TRACK_COLOR = "eef1f5"
CARD_BORDER_COLOR = "dde3ea"
CARD_TEXT_COLOR = "1f2937"
CARD_META_COLOR = "6b7280"


def _week_title(index: int, week: dict) -> str:
    title = f"Week {index:02d} Progress %"
    if week.get("label"):
        title += f"\n{week['label']}"
    return title


def _weighted(item: dict, week_id: str) -> float:
    percent = (item.get("entries") or {}).get(week_id, 0) or 0
    return (percent / 100) * (item.get("project_cost_percent") or 0)


def build_excel(project: dict, weeks: list[dict], items: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Tracker"

    n_static = len(STATIC_HEADERS)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    category_fill = PatternFill(start_color=CATEGORY_BG, end_color=CATEGORY_BG, fill_type="solid")
    category_font = Font(bold=True, color=CATEGORY_FONT_COLOR)
    subcategory_fill = PatternFill(start_color=SUBCATEGORY_BG, end_color=SUBCATEGORY_BG, fill_type="solid")
    subcategory_font = Font(bold=True, color=SUBCATEGORY_FONT_COLOR)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, title in enumerate(STATIC_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    col = n_static + 1
    for idx, week in enumerate(weeks, start=1):
        title_cell = ws.cell(row=1, column=col, value=_week_title(idx, week))
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = wrap_center
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(row=1, column=col + 1).fill = header_fill

        for offset, label in enumerate(("Progress %", "Weighted %")):
            sub_cell = ws.cell(row=2, column=col + offset, value=label)
            sub_cell.font = header_font
            sub_cell.fill = header_fill
            sub_cell.alignment = wrap_center
        col += 2

    row = 3
    cost_total = 0.0
    week_totals = [0.0] * len(weeks)

    for item in items:
        is_leaf = item.get("level") == LEVEL_ITEM
        code_cell = ws.cell(row=row, column=1, value=item.get("code") or "")
        code_cell.alignment = wrap_center
        desc_cell = ws.cell(row=row, column=2, value=item.get("description") or "")
        desc_cell.alignment = wrap_left

        if is_leaf:
            unit_cell = ws.cell(row=row, column=3, value=item.get("unit") or "")
            unit_cell.alignment = wrap_center
            ws.cell(row=row, column=4, value=item.get("suggested_quantity") or 0).alignment = wrap_center
            cost = item.get("project_cost_percent") or 0
            ws.cell(row=row, column=5, value=cost).alignment = wrap_center
            cost_total += cost

            col = n_static + 1
            for w_idx, week in enumerate(weeks):
                percent = (item.get("entries") or {}).get(week["id"], 0) or 0
                weighted = _weighted(item, week["id"])
                week_totals[w_idx] += weighted
                ws.cell(row=row, column=col, value=percent).alignment = wrap_center
                ws.cell(row=row, column=col + 1, value=round(weighted, 2)).alignment = wrap_center
                col += 2
        else:
            fill = category_fill if item.get("level") == LEVEL_CATEGORY else subcategory_fill
            font = category_font if item.get("level") == LEVEL_CATEGORY else subcategory_font
            desc_cell.font = font
            desc_cell.alignment = wrap_left
            code_cell.font = font
            last_col = n_static + len(weeks) * 2
            for c in range(1, last_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = fill
        row += 1

    total_row = row
    label_cell = ws.cell(row=total_row, column=1, value="TOTAL")
    label_cell.font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=5, value=round(cost_total, 2)).font = Font(bold=True)

    col = n_static + 1
    for w_idx in range(len(weeks)):
        ws.cell(row=total_row, column=col + 1, value=round(week_totals[w_idx], 2)).font = Font(bold=True)
        col += 2

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    col = n_static + 1
    for _ in weeks:
        ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions[get_column_letter(col + 1)].width = 12
        col += 2
    ws.row_dimensions[1].height = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---- Completion cards (Overall + per-category), mirroring progress_tracker.js ----


def _severity_for(percent: float) -> tuple[str, str]:
    if percent >= 80:
        return ("Nearly complete", GOOD_COLOR)
    if percent >= 40:
        return ("In progress", WARNING_COLOR)
    if percent > 0:
        return ("Just started", CRITICAL_COLOR)
    return ("Not started", NONE_COLOR)


def _group_label(header_item: dict, fallback: str) -> str:
    parts = [p for p in (header_item.get("code"), header_item.get("description")) if p]
    return " — ".join(parts) if parts else fallback


def _group_progress(children: list[dict], latest_week: dict | None) -> tuple[float, float, int]:
    total_cost = sum(child.get("project_cost_percent") or 0 for child in children)
    weighted_done = sum(_weighted(child, latest_week["id"]) for child in children) if latest_week else 0.0
    percent = (weighted_done / total_cost * 100) if total_cost > 0 else 0.0
    return percent, total_cost, len(children)


def _compute_category_groups(items: list[dict]) -> list[dict]:
    """Mirrors computeCardGroups in progress_tracker.js: each category's
    children are every leaf item up to the next category row (regardless of
    subcategory rows in between); each subcategory's children are just its
    own direct leaf items, up to the next subcategory/category row."""
    groups: list[dict] = []
    current = None
    current_sub = None
    for item in items:
        level = item.get("level")
        if level == LEVEL_CATEGORY:
            current = {"header": item, "children": [], "subgroups": []}
            groups.append(current)
            current_sub = None
        elif level == LEVEL_SUBCATEGORY:
            current_sub = {"header": item, "children": []}
            if current is not None:
                current["subgroups"].append(current_sub)
        elif level == LEVEL_ITEM:
            if current is not None:
                current["children"].append(item)
            if current_sub is not None:
                current_sub["children"].append(item)
    return groups


def _truncate_to_width(text: str, font_name: str, font_size: float, max_width: float) -> str:
    """Approximates the web card's CSS `text-overflow: ellipsis` for a
    single-line Paragraph, since reportlab has no equivalent — wrapping
    instead would make sibling cards' headers different heights and throw
    off the alignment of the value/meter/status rows beneath them."""
    from reportlab.pdfbase.pdfmetrics import stringWidth

    if stringWidth(text, font_name, font_size) <= max_width:
        return text
    ellipsis = "…"
    for end in range(len(text) - 1, 0, -1):
        candidate = text[:end].rstrip() + ellipsis
        if stringWidth(candidate, font_name, font_size) <= max_width:
            return candidate
    return ellipsis


def _build_meter(percent: float, color_hex: str, width: float, height: float = 6) -> Table:
    clamped = min(100.0, max(0.0, percent))
    width = max(width, 1.0)
    filled = width * (clamped / 100)
    empty = width - filled

    if filled <= 0:
        col_widths, bg_commands = [width], [("BACKGROUND", (0, 0), (0, 0), colors.HexColor(f"#{METER_TRACK_COLOR}"))]
    elif empty <= 0:
        col_widths, bg_commands = [width], [("BACKGROUND", (0, 0), (0, 0), colors.HexColor(f"#{color_hex}"))]
    else:
        col_widths = [filled, empty]
        bg_commands = [
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(f"#{color_hex}")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor(f"#{METER_TRACK_COLOR}")),
        ]

    meter = Table([[""] * len(col_widths)], colWidths=col_widths, rowHeights=[height])
    meter.setStyle(TableStyle(bg_commands + [
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return meter


def _build_overall_card(items: list[dict], weeks: list[dict], width: float) -> Table | None:
    leaf_items = [item for item in items if item.get("level") == LEVEL_ITEM]
    if not leaf_items:
        return None
    latest_week = weeks[-1] if weeks else None
    percent, _, _ = _group_progress(leaf_items, latest_week)
    clamped = min(100.0, max(0.0, percent))
    status_label, color_hex = _severity_for(percent)

    label_style = ParagraphStyle(
        "overall_label", fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor(f"#{CARD_META_COLOR}")
    )
    value_style = ParagraphStyle(
        "overall_value", fontName="Helvetica-Bold", fontSize=22, textColor=colors.HexColor(f"#{CARD_TEXT_COLOR}")
    )
    status_style = ParagraphStyle(
        "overall_status", fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor(f"#{color_hex}")
    )

    label_w = width * 0.22
    value_w = width * 0.13
    status_w = width * 0.17
    meter_w = width - label_w - value_w - status_w - 36

    row = Table(
        [[
            Paragraph("OVERALL PROJECT COMPLETION", label_style),
            Paragraph(f"{round(clamped)}%", value_style),
            _build_meter(percent, color_hex, meter_w, height=8),
            Paragraph(escape(status_label), status_style),
        ]],
        colWidths=[label_w, value_w, meter_w, status_w],
    )
    row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(f"#{CARD_BORDER_COLOR}")),
        ("LINEBEFORE", (0, 0), (0, -1), 3, colors.HexColor(f"#{color_hex}")),
    ]))
    return row


def _build_subcategory_row(sub: dict, latest_week: dict | None, card_width: float) -> Table:
    percent, total_cost, item_count = _group_progress(sub["children"], latest_week)
    clamped = min(100.0, max(0.0, percent))
    status_label, color_hex = _severity_for(percent)
    inner_width = card_width - 16

    name_style = ParagraphStyle(
        "subcat_name", fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.HexColor(f"#{CARD_TEXT_COLOR}")
    )
    value_style = ParagraphStyle(
        "subcat_value", fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.HexColor(f"#{color_hex}"), alignment=2
    )
    meta_style = ParagraphStyle(
        "subcat_meta", fontName="Helvetica", fontSize=6, textColor=colors.HexColor(f"#{CARD_META_COLOR}")
    )

    name_w = inner_width * 0.68
    value_w = inner_width - name_w
    title = _truncate_to_width(_group_label(sub["header"], "Untitled subcategory"), "Helvetica-Bold", 6.5, name_w - 2)
    item_word = "item" if item_count == 1 else "items"

    top_row = Table(
        [[Paragraph(escape(title), name_style), Paragraph(f"{round(clamped)}%", value_style)]],
        colWidths=[name_w, value_w],
    )
    top_row.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    block = Table(
        [
            [top_row],
            [_build_meter(percent, color_hex, inner_width, height=3)],
            [Paragraph(f"{item_count} {item_word} &middot; {total_cost:.2f}% of cost", meta_style)],
        ],
        colWidths=[inner_width],
    )
    block.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor(f"#{CARD_BORDER_COLOR}")),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
    ]))
    return block


def _build_category_card(group: dict, latest_week: dict | None, card_width: float) -> Table:
    percent, total_cost, item_count = _group_progress(group["children"], latest_week)
    clamped = min(100.0, max(0.0, percent))
    status_label, color_hex = _severity_for(percent)

    header_style = ParagraphStyle(
        "card_header", fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor(f"#{CARD_TEXT_COLOR}")
    )
    value_style = ParagraphStyle(
        "card_value", fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor(f"#{CARD_TEXT_COLOR}")
    )
    status_style = ParagraphStyle(
        "card_status", fontName="Helvetica-Bold", fontSize=7, textColor=colors.HexColor(f"#{color_hex}")
    )
    meta_style = ParagraphStyle(
        "card_meta", fontName="Helvetica", fontSize=6.5, textColor=colors.HexColor(f"#{CARD_META_COLOR}")
    )

    inner_width = card_width - 16
    # Truncated to a single line (matching the web card's ellipsis rule) so
    # every card's header takes the same height — otherwise a longer title
    # wrapping to a 2nd line pushes that card's value/meter/status down
    # relative to its neighbors in the same grid row.
    title = _truncate_to_width(
        _group_label(group["header"], "Untitled category"), "Helvetica-Bold", 8, inner_width
    )
    meter = _build_meter(percent, color_hex, inner_width, height=5)
    item_word = "item" if item_count == 1 else "items"
    meta_text = f"{item_count} {item_word} &middot; {total_cost:.2f}% of cost"

    rows = [
        [Paragraph(escape(title), header_style)],
        [Paragraph(f"{round(clamped)}%", value_style)],
        [meter],
        [Paragraph(escape(status_label), status_style)],
        [Paragraph(meta_text, meta_style)],
    ]
    for sub in group["subgroups"]:
        rows.append([_build_subcategory_row(sub, latest_week, card_width)])

    card = Table(rows, colWidths=[card_width])
    card.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(f"#{CARD_BORDER_COLOR}")),
    ]))
    return card


def _build_category_cards_grid(items: list[dict], weeks: list[dict], page_width: float) -> Table | None:
    groups = _compute_category_groups(items)
    if not groups:
        return None
    latest_week = weeks[-1] if weeks else None

    # 3 per row, not 4 — cards now carry a subcategory breakdown, so the
    # extra width keeps the subcategory name/percent/meta lines from
    # feeling cramped.
    cols = 3
    gap = 8
    card_width = (page_width - gap * (cols - 1)) / cols

    rows: list[list] = []
    current_row: list = []
    for group in groups:
        current_row.append(_build_category_card(group, latest_week, card_width))
        if len(current_row) == cols:
            rows.append(current_row)
            current_row = []
    if current_row:
        current_row.extend([""] * (cols - len(current_row)))
        rows.append(current_row)

    grid = Table(rows, colWidths=[card_width] * cols)
    grid.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), gap),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), gap),
    ]))
    return grid


def build_pdf(project: dict, weeks: list[dict], items: list[dict]) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18
    )
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>{escape(project.get('name', ''))} — Progress Tracker</b>", styles["Title"])

    page_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin

    flowables: list = [title, Spacer(1, 10)]
    overall_card = _build_overall_card(items, weeks, page_width)
    if overall_card is not None:
        flowables.append(overall_card)
        flowables.append(Spacer(1, 12))
    category_cards = _build_category_cards_grid(items, weeks, page_width)
    if category_cards is not None:
        flowables.append(category_cards)
        flowables.append(Spacer(1, 12))

    header_style = ParagraphStyle(
        "header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        alignment=1,
        textColor=colors.HexColor(f"#{HEADER_FONT_COLOR}"),
    )
    desc_style = ParagraphStyle(
        "desc", parent=styles["Normal"], fontName="Helvetica", fontSize=7, leading=9
    )
    desc_category_style = ParagraphStyle(
        "desc_category",
        parent=desc_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{CATEGORY_FONT_COLOR}"),
    )
    desc_subcategory_style = ParagraphStyle(
        "desc_subcategory",
        parent=desc_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{SUBCATEGORY_FONT_COLOR}"),
    )

    def header_paragraph(text: str) -> Paragraph:
        return Paragraph(escape(text).replace("\n", "<br/>"), header_style)

    n_static = len(STATIC_HEADERS)
    header_row1 = [header_paragraph(h) for h in STATIC_HEADERS]
    header_row2 = [""] * n_static
    for idx, week in enumerate(weeks, start=1):
        header_row1.extend([header_paragraph(_week_title(idx, week)), ""])
        header_row2.extend([header_paragraph("Progress %"), header_paragraph("Weighted %")])

    data = [header_row1, header_row2]
    span_commands = [("SPAN", (c, 0), (c, 1)) for c in range(n_static)]
    col = n_static
    for _ in weeks:
        span_commands.append(("SPAN", (col, 0), (col + 1, 0)))
        col += 2

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor(f"#{HEADER_BG}")),
        ("TEXTCOLOR", (0, 0), (-1, 1), colors.HexColor(f"#{HEADER_FONT_COLOR}")),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde3ea")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
    ]
    style_commands.extend(span_commands)

    cost_total = 0.0
    week_totals = [0.0] * len(weeks)

    for r, item in enumerate(items, start=2):
        is_leaf = item.get("level") == LEVEL_ITEM
        description = escape(item.get("description") or "")

        if is_leaf:
            row = [item.get("code") or "", Paragraph(description, desc_style)]
            cost = item.get("project_cost_percent") or 0
            row += [item.get("unit") or "", item.get("suggested_quantity") or 0, cost]
            cost_total += cost
            for w_idx, week in enumerate(weeks):
                percent = (item.get("entries") or {}).get(week["id"], 0) or 0
                weighted = _weighted(item, week["id"])
                week_totals[w_idx] += weighted
                row += [percent, round(weighted, 2)]
        else:
            if item.get("level") == LEVEL_CATEGORY:
                bg = colors.HexColor(f"#{CATEGORY_BG}")
                desc_para = Paragraph(description, desc_category_style)
            else:
                bg = colors.HexColor(f"#{SUBCATEGORY_BG}")
                desc_para = Paragraph(description, desc_subcategory_style)
            row = [item.get("code") or "", desc_para] + ["", "", ""] + ["", ""] * len(weeks)
            fg = desc_para.style.textColor
            style_commands.append(("BACKGROUND", (0, r), (-1, r), bg))
            style_commands.append(("TEXTCOLOR", (0, r), (0, r), fg))
            style_commands.append(("FONTNAME", (0, r), (0, r), "Helvetica-Bold"))

        data.append(row)

    total_row_idx = len(data)
    total_row = ["TOTAL", "", "", "", round(cost_total, 2)]
    for w_idx in range(len(weeks)):
        total_row += ["", round(week_totals[w_idx], 2)]
    data.append(total_row)
    style_commands.append(("SPAN", (0, total_row_idx), (3, total_row_idx)))
    style_commands.append(("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"))

    fixed_widths = [page_width * f for f in (0.05, 0.28, 0.05, 0.07, 0.07)]
    remaining = page_width - sum(fixed_widths)
    week_col_count = max(1, len(weeks) * 2)
    col_widths = fixed_widths + [remaining / week_col_count] * (len(weeks) * 2)

    table = Table(data, repeatRows=2, colWidths=col_widths)
    table.setStyle(TableStyle(style_commands))
    flowables.append(table)

    doc.build(flowables)
    buf.seek(0)
    return buf
